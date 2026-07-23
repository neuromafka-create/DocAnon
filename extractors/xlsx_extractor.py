from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.models import ExtractionResult
from core.xlsx_cells import cell_value_to_str
from extractors.base import TextExtractor


class XlsxExtractor(TextExtractor):
    """Извлечение текста из Excel (.xlsx / .xlsm).

    Каждый лист → блок с заголовком; строки таблиц → " | "-разделённые ячейки,
    чтобы Presidio/spaCy видели контекст (ФИО рядом с «Контакт», ИНН в колонке и т.д.).
    """

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm"]

    def extract(self, file_path: Path) -> ExtractionResult:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        sheets_text: list[str] = []
        total_rows = 0

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines: list[str] = [f"=== {sheet_name} ==="]
                sheet_rows = 0

                for row in ws.iter_rows(values_only=True):
                    cells = [cell_value_to_str(c) for c in row]
                    cells = [c for c in cells if c]
                    if not cells:
                        continue
                    lines.append(" | ".join(cells))
                    sheet_rows += 1

                if sheet_rows > 0:
                    sheets_text.append("\n".join(lines))
                    total_rows += sheet_rows
        finally:
            wb.close()

        text = "\n\n".join(sheets_text)
        return ExtractionResult(
            text=text,
            source_format="xlsx",
            pages=sheets_text,
            metadata={
                "file_name": file_path.name,
                "sheets": len(sheets_text),
                "rows": total_rows,
            },
        )
