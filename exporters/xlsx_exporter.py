from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

from core.models import AnonymizedDocument
from core.xlsx_cells import apply_mapping_to_cell_value
from exporters.base import Exporter

XLSX_EXTENSIONS = {".xlsx", ".xlsm"}


class XlsxExporter(Exporter):
    """Round-trip: пишет обезличенные значения обратно в ячейки .xlsx/.xlsm.

    Структура книги, стили, формулы и merged-ячейки сохраняются.
    """

    def export(self, doc: AnonymizedDocument, output_path: Path) -> Path:
        if not doc.source_path:
            raise ValueError(
                "Для экспорта XLSX нужен source_path в AnonymizedDocument"
            )
        source = Path(doc.source_path)
        if source.suffix.lower() not in XLSX_EXTENSIONS:
            raise ValueError(f"Исходный файл не Excel: {source.suffix}")
        if not source.exists():
            raise FileNotFoundError(f"Исходный файл не найден: {source}")
        return self.export_workbook(source, doc.mapping, output_path)

    def export_workbook(
        self,
        source_path: Path,
        mapping: dict[str, str],
        output_path: Path,
    ) -> Path:
        source_path = Path(source_path)
        suffix = source_path.suffix.lower()
        if suffix not in XLSX_EXTENSIONS:
            raise ValueError(f"Неподдерживаемый формат: {suffix}")

        output_path = Path(output_path)
        if output_path.suffix.lower() not in XLSX_EXTENSIONS:
            output_path = output_path.with_suffix(suffix)

        keep_vba = suffix == ".xlsm"
        wb = load_workbook(str(source_path), data_only=False, keep_vba=keep_vba)

        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        if not isinstance(cell, Cell):
                            continue
                        cell.value = apply_mapping_to_cell_value(
                            cell.value, mapping
                        )
            wb.save(str(output_path))
        finally:
            wb.close()

        return output_path
