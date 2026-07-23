from __future__ import annotations

"""Обогащение mapping для Excel: ячейки ≠ flattened-строка с « | ».

Проблема: extract склеивает ``ООО | ФИО``, NER даёт один span, а в файле
это две ячейки — ФИО не попадает в mapping и не маскируется при export.
"""

import logging
import re
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from core.models import Entity
from core.xlsx_cells import cell_value_to_str

if TYPE_CHECKING:
    from core.mapping import MappingStore
    from core.pipeline import AnonymizationPipeline

logger = logging.getLogger(__name__)

XLSX_SUFFIXES = {".xlsx", ".xlsm"}
_PIPE = " | "

_HEADER_WORDS = frozenset({
    "компания", "контакт", "клиент", "дата", "сумма", "долг",
    "фио", "имя", "название", "организация", "телефон", "email",
    "e-mail", "адрес", "инн", "примечание", "комментарий",
})


def is_xlsx_path(path: Path) -> bool:
    return path.suffix.lower() in XLSX_SUFFIXES


def _is_junk_surface(text: str) -> bool:
    """Служебные/битые фрагменты extract, не ПДн."""
    t = text.strip()
    if len(t) < 2:
        return True
    if "===" in t:
        return True
    if "\n" in t:
        return True
    if t.casefold() in _HEADER_WORDS:
        return True
    return False


def iter_xlsx_cell_texts(path: Path) -> list[str]:
    """Все непустые строковые значения ячеек (data_only)."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    values: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    text = cell_value_to_str(cell)
                    if text:
                        values.append(text)
    finally:
        wb.close()
    return values


def _looks_like_org(text: str) -> bool:
    t = text.strip().upper()
    return bool(
        re.match(
            r"^(ООО|ОАО|АО|ЗАО|ПАО|ИП|НКО|ФГУП|ГУП)\b",
            t,
        )
        or "«" in text
        or '"' in text
    )


def _register_surface(
    store: MappingStore,
    surface: str,
    label: str,
    source_file: str,
) -> None:
    surface = surface.strip()
    if _is_junk_surface(surface):
        return
    store.register(surface, label, source_file=source_file)


def _split_pipe_fragments(text: str) -> list[str]:
    if _PIPE not in text:
        return [text] if text.strip() else []
    return [p.strip() for p in text.split(_PIPE) if p.strip()]


def enrich_store_from_xlsx(
    path: Path,
    pipeline: AnonymizationPipeline,
    store: MappingStore,
    *,
    entities: list[Entity] | None = None,
    source_file: str = "",
) -> list[Entity]:
    """Дополнить MappingStore поверхностями из ячеек + анализ каждой ячейки.

    Returns:
        Дополнительные entity (cell-level), для статистики/preview.
    """
    path = Path(path)
    if not is_xlsx_path(path) or not path.exists():
        return []

    source_file = source_file or path.name
    extra: list[Entity] = []

    # 1) Разобрать spans с « | » из полного анализа
    if entities:
        for ent in entities:
            parts = _split_pipe_fragments(ent.text)
            if len(parts) <= 1:
                continue
            for part in parts:
                part_ents = pipeline.analyze(part)
                if part_ents:
                    for pe in part_ents:
                        _register_surface(store, pe.text, pe.label, source_file)
                        _register_surface(store, part, pe.label, source_file)
                        extra.append(pe)
                else:
                    label = (
                        "ORGANIZATION"
                        if _looks_like_org(part)
                        else (
                            "PERSON"
                            if ent.label == "PERSON"
                            or (part[:1].isupper() and " " in part)
                            else ent.label
                        )
                    )
                    _register_surface(store, part, label, source_file)

    # 2) Каждая ячейка — отдельный мини-документ
    cells = iter_xlsx_cell_texts(path)
    logger.info("XLSX enrich: %d cells in %s", len(cells), path.name)

    for cell_text in cells:
        if _is_junk_surface(cell_text):
            continue
        cell_ents = pipeline.analyze(cell_text)
        if not cell_ents:
            continue
        for pe in cell_ents:
            if _is_junk_surface(pe.text):
                continue
            _register_surface(store, pe.text, pe.label, source_file)
            extra.append(pe)
        # целая ячейка → placeholder доминирующей сущности
        best = max(cell_ents, key=lambda e: (len(e.text), e.confidence))
        if _is_junk_surface(best.text):
            continue
        if (
            best.text in cell_text
            or cell_text in best.text
            or _norm(best.text) == _norm(cell_text)
        ):
            _register_surface(store, cell_text, best.label, source_file)

    return extra


def enrich_dict_mapping_from_xlsx(
    path: Path,
    pipeline: AnonymizationPipeline,
    mapping: dict[str, str],
    *,
    entities: list[Entity] | None = None,
) -> dict[str, str]:
    """То же для single-file mapping (generic placeholders)."""
    from core.mapping import MappingStore

    store = MappingStore(person_canonical_mapping=True)
    # seed from existing mapping — register with labels from placeholders
    for original, ph in mapping.items():
        label = _label_from_placeholder(ph)
        store.register(original, label)

    enrich_store_from_xlsx(
        path, pipeline, store, entities=entities, source_file=path.name
    )

    # merge: keep original placeholders where possible, add new surfaces
    # with same style as single-file (Entity.placeholder) when new
    from core.models import ENTITY_PLACEHOLDERS

    out = dict(mapping)
    for original, ph in store.to_mapping_dict().items():
        if original in out:
            continue
        # new surface: use generic type placeholder for single-file consistency
        label = _label_from_placeholder(ph)
        out[original] = ENTITY_PLACEHOLDERS.get(label, ph)
    return out


def _label_from_placeholder(ph: str) -> str:
    inner = ph.strip("<>")
    if "_" in inner and inner.rsplit("_", 1)[-1].isdigit():
        return inner.rsplit("_", 1)[0]
    # generic <PERSON> / <ORG>
    reverse = {
        "PERSON": "PERSON",
        "ORG": "ORGANIZATION",
        "LOC": "LOCATION",
        "INN": "RU_INN",
        "SNILS": "RU_SNILS",
        "PHONE": "PHONE_NUMBER",
        "EMAIL": "EMAIL_ADDRESS",
        "PASSPORT_RF": "RU_PASSPORT",
        "CAR_PLATE": "RU_VEHICLE_PLATE",
        "DRIVER_LICENSE": "RU_DRIVER_LICENSE",
        "ACCOUNT": "RU_ACCOUNT",
        "BIK": "RU_BIK",
        "BANK_CARD": "CREDIT_CARD",
        "IP": "IP_ADDRESS",
        "MAC": "MAC_ADDRESS",
        "GEO": "GPS_COORDS",
        "TG_CHAT_ID": "TG_CHAT_ID",
    }
    return reverse.get(inner, inner)


def _norm(text: str) -> str:
    return " ".join(text.strip().split()).casefold()
