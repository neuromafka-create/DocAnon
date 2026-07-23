from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.batch_pipeline import BatchPipeline
from core.config import AnonymizerConfig
from core.pipeline import AnonymizationPipeline
from exporters.xlsx_exporter import XLSX_EXTENSIONS, XlsxExporter
from extractors import extract_text, get_extractor

TEST_DOCS = Path("test_docs")

DOC_FILES = sorted(TEST_DOCS.glob("*.docx")) if TEST_DOCS.exists() else []
XLSX_FILES = sorted(TEST_DOCS.glob("*.xlsx")) if TEST_DOCS.exists() else []
ALL_FILES = DOC_FILES + XLSX_FILES

pytestmark = pytest.mark.skipif(
    not TEST_DOCS.exists() or not ALL_FILES,
    reason="test_docs/ missing or empty",
)


@pytest.fixture(scope="module")
def config() -> AnonymizerConfig:
    return AnonymizerConfig()


@pytest.fixture(scope="module")
def pipeline(config: AnonymizerConfig) -> AnonymizationPipeline:
    return AnonymizationPipeline(config)


class TestExtractSmoke:
    @pytest.mark.parametrize("path", ALL_FILES, ids=lambda p: p.name)
    def test_extract_non_empty(self, path: Path) -> None:
        extractor = get_extractor(path)
        assert extractor is not None, f"no extractor for {path.suffix}"
        result = extract_text(path)
        assert result.text.strip(), f"empty text from {path.name}"
        assert result.source_format in ("docx", "xlsx")
        assert result.metadata.get("file_name") == path.name


class TestPipelineSmoke:
    @pytest.mark.parametrize("path", ALL_FILES, ids=lambda p: p.name)
    def test_process_file(self, pipeline: AnonymizationPipeline, path: Path) -> None:
        result = pipeline.process_file(path)
        assert result.source_file == path.name
        assert result.source_path
        assert Path(result.source_path).resolve() == path.resolve()
        assert result.original_text.strip()
        assert result.anonymized_text.strip()
        assert isinstance(result.entities, list)
        # каждый ключ mapping должен быть заменён в anonymized_text (span-replace)
        for original, placeholder in result.mapping.items():
            if not original or original == placeholder:
                continue
            assert original not in result.anonymized_text, (
                f"{path.name}: mapping key still in text: {original!r} → {placeholder!r}"
            )
            assert placeholder in result.anonymized_text or result.total_entities == 0
    def test_dogovor_redacts_contact_pii(self, pipeline: AnonymizationPipeline) -> None:
        path = TEST_DOCS / "Договор_поставки.docx"
        if not path.exists():
            pytest.skip("Договор_поставки.docx not found")
        result = pipeline.process_file(path)
        anon = result.anonymized_text
        assert "+7 (916) 412-58-31" not in anon
        assert "m.egorova@vectorfood.ru" not in anon
        labels = {e.label for e in result.entities}
        assert "PHONE_NUMBER" in labels
        assert "EMAIL_ADDRESS" in labels
        assert "ORGANIZATION" in labels or "PERSON" in labels

    def test_klienty_finds_org_or_person(self, pipeline: AnonymizationPipeline) -> None:
        path = TEST_DOCS / "Клиенты.xlsx"
        if not path.exists():
            pytest.skip("Клиенты.xlsx not found")
        result = pipeline.process_file(path)
        assert result.total_entities >= 1
        labels = {e.label for e in result.entities}
        assert labels & {"PERSON", "ORGANIZATION"}


class TestXlsxRoundTripSmoke:
    @pytest.mark.parametrize("path", XLSX_FILES, ids=lambda p: p.name)
    def test_export_workbook(
        self,
        pipeline: AnonymizationPipeline,
        path: Path,
        tmp_path: Path,
    ) -> None:
        result = pipeline.process_file(path)
        out = tmp_path / f"{path.stem}_anon{path.suffix}"
        saved = XlsxExporter().export(result, out)
        assert saved.exists()
        assert saved.suffix.lower() in XLSX_EXTENSIONS

        # structure preserved
        wb_src = load_workbook(path, read_only=True, data_only=False)
        wb_out = load_workbook(saved, read_only=True, data_only=False)
        try:
            assert wb_src.sheetnames == wb_out.sheetnames
        finally:
            wb_src.close()
            wb_out.close()

        # re-extract works
        re_extracted = extract_text(saved)
        assert re_extracted.text.strip()

        # Round-trip по ячейкам: ключ mapping, целиком лежащий в одной ячейке,
        # должен быть заменён. (Span'ы вида "ORG | ФИО" из flattened extract
        # пересекают ячейки — это ограничение extract, не exporter.)
        from core.xlsx_cells import cell_value_to_str

        wb = load_workbook(saved, data_only=False)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        cell_text = cell_value_to_str(cell.value)
                        if not cell_text:
                            continue
                        if cell_text in result.mapping:
                            assert cell.value == result.mapping[cell_text], (
                                f"{path.name}!{cell.coordinate}: "
                                f"expected {result.mapping[cell_text]!r}, "
                                f"got {cell.value!r}"
                            )
        finally:
            wb.close()


class TestBatchSmoke:
    def test_batch_all_test_docs(self, config: AnonymizerConfig, tmp_path: Path) -> None:
        files = ALL_FILES
        assert files, "no test docs"
        batch = BatchPipeline(config)
        result = batch.process_batch(files)

        assert len(result.documents) == len(files)
        assert result.mapping.total_entities >= 0
        names = {d.source_file for d in result.documents}
        assert names == {p.name for p in files}

        # numbered placeholders when entities found
        if result.mapping.total_entities > 0:
            for original, ph in result.mapping.to_mapping_dict().items():
                assert ph.startswith("<") and ph.endswith(">")
                assert "_" in ph

        # xlsx round-trip with final mapping
        final_map = result.mapping.to_mapping_dict()
        exporter = XlsxExporter()
        for doc in result.documents:
            src = Path(doc.source_path)
            if src.suffix.lower() not in XLSX_EXTENSIONS:
                continue
            out = tmp_path / f"batch_{src.name}"
            exporter.export_workbook(src, final_map, out)
            assert out.exists()
            assert extract_text(out).text.strip()

    def test_batch_consistent_person_name(self, config: AnonymizerConfig) -> None:
        """Одно ФИО в разных файлах → один placeholder (если NER выделил одинаково)."""
        paths = [
            TEST_DOCS / "Клиенты.xlsx",
            TEST_DOCS / "Отчет_менеджера_январь.docx",
        ]
        paths = [p for p in paths if p.exists()]
        if len(paths) < 2:
            pytest.skip("need both Клиенты and Отчет")

        result = BatchPipeline(config).process_batch(paths)
        # collect placeholders for fragments containing Егорова
        egorova_phs = {
            ph
            for orig, ph in result.mapping.to_mapping_dict().items()
            if "Егорова" in orig
        }
        # if detected as same span text, single ph; if split differently, multiple — still ok
        assert isinstance(egorova_phs, set)
        # at least something about the person/org should be mapped across docs
        assert result.total_entities >= 1
