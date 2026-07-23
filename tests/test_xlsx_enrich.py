from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from core.batch_pipeline import BatchPipeline
from core.config import AnonymizerConfig
from core.mapping import MappingStore
from core.pipeline import AnonymizationPipeline
from core.xlsx_enrich import enrich_store_from_xlsx, iter_xlsx_cell_texts
from exporters.xlsx_exporter import XlsxExporter


@pytest.fixture
def clients_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "Клиенты.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["Компания", "Контакт"])
    ws.append(["ООО «ВекторФуд»", "Егорова Марина Викторовна"])
    ws.append(["ООО «ГастроМаркет»", "Новиков Павел Игоревич"])
    wb.save(path)
    return path


class TestXlsxEnrich:
    def test_iter_cells(self, clients_xlsx: Path) -> None:
        cells = iter_xlsx_cell_texts(clients_xlsx)
        assert "Новиков Павел Игоревич" in cells
        assert "ООО «ГастроМаркет»" in cells

    def test_enrich_registers_person_cell(self, clients_xlsx: Path) -> None:
        pipe = AnonymizationPipeline(AnonymizerConfig())
        store = MappingStore()
        text = (
            "=== Клиенты ===\n"
            "Компания | Контакт\n"
            "ООО «ВекторФуд» | Егорова Марина Викторовна\n"
            "ООО «ГастроМаркет» | Новиков Павел Игоревич"
        )
        entities = pipe.analyze(text)
        pipe.register_entities(text, entities, store, source_file=clients_xlsx.name)
        enrich_store_from_xlsx(
            clients_xlsx, pipe, store, entities=entities, source_file=clients_xlsx.name
        )

        # ФИО из отдельной ячейки должно быть в mapping
        m = store.to_mapping_dict()
        assert any("Новиков" in k for k in m), f"keys={list(m)}"
        # либо точное, либо канон-алиас
        assert (
            "Новиков Павел Игоревич" in m
            or store.get_placeholder("Новиков Павел Игоревич") is not None
        )

    def test_batch_export_masks_both_contacts(
        self, clients_xlsx: Path, tmp_path: Path
    ) -> None:
        batch = BatchPipeline(pipeline=AnonymizationPipeline(AnonymizerConfig()))
        result = batch.process_batch([clients_xlsx])
        out = tmp_path / "anon.xlsx"
        XlsxExporter().export_workbook(
            clients_xlsx, result.mapping.to_mapping_dict(), out
        )

        wb = load_workbook(out, data_only=True)
        ws = wb.active
        # B2 and B3 — контакты
        b2 = str(ws["B2"].value or "")
        b3 = str(ws["B3"].value or "")
        wb.close()

        assert "Егорова" not in b2, b2
        assert "Новиков" not in b3, b3
        assert "<" in b2 and "<" in b3

    def test_single_file_export_masks_novikov(
        self, clients_xlsx: Path, tmp_path: Path
    ) -> None:
        pipe = AnonymizationPipeline(AnonymizerConfig())
        doc = pipe.process_file(clients_xlsx)
        out = tmp_path / "single.xlsx"
        XlsxExporter().export(doc, out)

        wb = load_workbook(out, data_only=True)
        ws = wb.active
        b3 = str(ws["B3"].value or "")
        wb.close()
        assert "Новиков" not in b3, f"mapping keys sample={list(doc.mapping)[:10]}, B3={b3}"
        assert "<" in b3
