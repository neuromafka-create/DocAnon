from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from core.models import AnonymizedDocument
from core.xlsx_cells import (
    apply_mapping_to_cell_value,
    apply_mapping_to_text,
    cell_value_to_str,
)
from exporters.xlsx_exporter import XlsxExporter


class TestCellHelpers:
    def test_cell_value_to_str_int_float(self) -> None:
        assert cell_value_to_str(7707083893) == "7707083893"
        assert cell_value_to_str(7707083893.0) == "7707083893"
        assert cell_value_to_str(1.5) == "1.5"

    def test_apply_mapping_longest_first(self) -> None:
        mapping = {
            "Иванов": "<PERSON_2>",
            "Иванов Иван": "<PERSON_1>",
        }
        assert apply_mapping_to_text("Иванов Иван", mapping) == "<PERSON_1>"

    def test_apply_mapping_to_numeric_cell(self) -> None:
        mapping = {"7707083893": "<RU_INN_1>"}
        assert apply_mapping_to_cell_value(7707083893, mapping) == "<RU_INN_1>"
        assert apply_mapping_to_cell_value(7707083893.0, mapping) == "<RU_INN_1>"

    def test_formulas_untouched(self) -> None:
        mapping = {"A1": "<X>"}
        assert apply_mapping_to_cell_value("=A1+1", mapping) == "=A1+1"

    def test_partial_string_replace(self) -> None:
        mapping = {"Егорова Марина Викторовна": "<PERSON_1>"}
        value = "Контакт: Егорова Марина Викторовна"
        assert (
            apply_mapping_to_cell_value(value, mapping)
            == "Контакт: <PERSON_1>"
        )


class TestXlsxExporter:
    @pytest.fixture
    def sample_xlsx(self, tmp_path: Path) -> Path:
        path = tmp_path / "clients.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Клиенты"
        ws["A1"] = "Компания"
        ws["B1"] = "Контакт"
        ws["C1"] = "ИНН"
        ws["A2"] = "ООО «ВекторФуд»"
        ws["B2"] = "Егорова Марина Викторовна"
        ws["C2"] = 7707083893
        ws["D2"] = "=C2"  # formula must survive
        wb.save(path)
        return path

    def test_export_workbook_replaces_cells(
        self, sample_xlsx: Path, tmp_path: Path
    ) -> None:
        mapping = {
            "Егорова Марина Викторовна": "<PERSON_1>",
            "7707083893": "<RU_INN_1>",
            "ООО «ВекторФуд»": "<ORG_1>",
        }
        out = tmp_path / "clients_anonymized.xlsx"
        result = XlsxExporter().export_workbook(sample_xlsx, mapping, out)
        assert result == out
        assert out.exists()

        wb = load_workbook(out, data_only=False)
        ws = wb["Клиенты"]
        assert ws["A1"].value == "Компания"
        assert ws["B2"].value == "<PERSON_1>"
        assert ws["C2"].value == "<RU_INN_1>"
        assert ws["A2"].value == "<ORG_1>"
        assert ws["D2"].value == "=C2"
        wb.close()

    def test_export_from_document(
        self, sample_xlsx: Path, tmp_path: Path
    ) -> None:
        doc = AnonymizedDocument(
            original_text="…",
            anonymized_text="…",
            mapping={
                "Егорова Марина Викторовна": "<PERSON>",
                "7707083893": "<INN>",
            },
            source_file=sample_xlsx.name,
            source_path=str(sample_xlsx),
        )
        out = tmp_path / "from_doc.xlsx"
        path = XlsxExporter().export(doc, out)
        wb = load_workbook(path)
        assert wb.active["B2"].value == "<PERSON>"
        assert wb.active["C2"].value == "<INN>"
        wb.close()

    def test_export_requires_source_path(self, tmp_path: Path) -> None:
        doc = AnonymizedDocument(
            original_text="",
            anonymized_text="",
            mapping={},
        )
        with pytest.raises(ValueError, match="source_path"):
            XlsxExporter().export(doc, tmp_path / "x.xlsx")

    def test_roundtrip_with_extractor(
        self, sample_xlsx: Path, tmp_path: Path
    ) -> None:
        from extractors import extract_text

        mapping = {
            "Егорова Марина Викторовна": "<PERSON_1>",
            "7707083893": "<RU_INN_1>",
        }
        out = tmp_path / "anon.xlsx"
        XlsxExporter().export_workbook(sample_xlsx, mapping, out)

        extracted = extract_text(out)
        assert "Егорова" not in extracted.text
        assert "7707083893" not in extracted.text
        assert "<PERSON_1>" in extracted.text
        assert "<RU_INN_1>" in extracted.text
